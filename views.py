from django.http.response import HttpResponse

__author__ = 'eoo'

from openpyxl import load_workbook


def map_show(response):
    #https://maps.googleapis.com/maps/api/directions/json?origin=Toronto&destination=Montreal&key=API_KEY
    pass
def results(response):
    wb2 = load_workbook('gitoc.xlsx')
    #a = "Milan"
    #b = "Italy"
    a = response.GET.get("start_location");
    b = response.GET.get("end_location");

    #seeet_names = (wb2.get_sheet_names())
    c = wb2.get_sheet_by_name("Prices");
    html = "";


    #for row in c.iter_rows(row_offset=1):
    #    html += row
    #n = c.cell(row=1)
    n = "";
    i = 1;
    x = []
    for i in range(i,c.get_highest_row()):
        if(str(c.cell(row=i, column=1).value) == a):
            if(str(c.cell(row=i,column=2).value) == b):
                n+= str(c.cell(row=i,column=10).value) + "<br/>";

    #for row in c.iter_rows('A1:C2'):
    #    for cell in row:
    #        html += str(cell['']);

    #for row in c.iter_rows():
    #    html += row['Start']
        #print(row)
        #if(row)
        #/html += str(row) + "\n"

    #
    #    for cell in row:
    #        #print(cell.value)

    return HttpResponse(str(n));
